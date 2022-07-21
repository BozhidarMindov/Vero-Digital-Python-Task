import pandas
import requests
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import math
import argparse
import time


API = "https://api.baubuddy.de/dev/index.php/v1/vehicles/select/active"
COLOR_API = "https://api.baubuddy.de/dev/index.php/v1/labels/"


def main():
    #uncomment here an in the bottom of file in order to see the program's execution time
    start_time = time.time()

    #making a request to the API for data
    response = requests.get(API).json()

    #filtering API responses, so that only the json items that have an "hu" value are processed
    filtered_response = []
    for item in response:
        if item["hu"]:
            filtered_response.append(item)

    #mock object that is purely for testing
    filtered_response.append(
    {
        "rnr": "21",
        "gruppe": "Lieferwagen",
        "kurzname": "PB-XY 406 (PKW Hänger)",
        "langtext": "PKW Anhänger mit Alubordwänden",
        "info": "2.2 to Zuladung",
        "sort": "0",
        "lagerort": "Paderborn",
        "lteartikel": "",
        "businessUnit": "Gerüstbau",
        "vondat": "2016-10-01",
        "bisdat": "2016-10-01",
        "hu": "2017-03-01",
        "asu": "2016-10-01",
        "createdOn": "2016-10-01",
        "editedOn": "2022-01-11T10:52:06Z",
        "fuelConsumption": "0.0",
        "priceInformation": "0.0",
        "safetyCheckDate": "2016-10-01",
        "tachographTestDate": "2016-10-01",
        "gb1": "Gerüstbau",
        "ownerId": "85",
        "userId": "85",
        "externalId": "",
        "vin": "",
        "labelIds": "76",
        "bleGroupEnum": "",
        "profile_picture": "https://api.baubuddy.de/branches/api-develop/infomaterial/Dokumente_vero_test/RNR_21/Bilder/IMG_20171124_13304949851817.jpg",
        "thumbPathUrl": "https://api.baubuddy.de/branches/api-develop/infomaterial/Dokumente_vero_test/RNR_21/Bilder/Thumbs/IMG_20171124_13304949851817.jpg"
      }
    )


    #since the keys in every json item in the response are the same, we can save their names and use them later to check for valid user input
    keys = [key.lower() for key in filtered_response[0]]


    #asking for the names of additional columns to be printed
    parser = argparse.ArgumentParser(description="Enter as many keys as you would like")

    parser.add_argument('-k', '--keys', type=str, nargs="+", required=True, help = "Keys")
    parser.add_argument('-c', '--colored', default = "True")

    args = parser.parse_args()

    #-----------------Old way------------
    # k = args.keys.lower()
    # print(k)
    # additional_column_names = k.split(" ")


    #------------------New Way-----------------
    #allows us to pass an arbitrary number of strings
    k = args.keys
    updated_k = " ".join(k).lower()
    additional_column_names = updated_k.split(" ")
    print(additional_column_names)

    #if the user wants all columns printed, they could just use keywword "all"
    if "all" in additional_column_names:
        additional_column_names = keys

    #if the user has entered some invalid input the program breaks
    for name in additional_column_names:
        if name not in keys:
            print("You have entered invalid input!")
            exit()


    #converting from string to bool, because python has a problerm dealing with command line booleans (and booleans in general)
    c = args.colored.lower()
    if c == "false":
        c = False
    else:
        c = True


    #creating a list of columns that would later be placed in the xlsx file
    columns_to_print = []
    if "rnr" in additional_column_names:
        pass
    else:
        columns_to_print.append("rnr")

    if "gruppe" in additional_column_names:
        pass
    else:
        columns_to_print.append("gruppe")

    #basically sorts the columns that the user requested in teh right order
    for item in filtered_response[0]:
        if item.lower() in additional_column_names:
            columns_to_print.append(item)


    #getting the current date and iso-formating it
    today = datetime.now()
    current_date_iso_formatted = today.isoformat().split("T").pop(0)


    #placing the data we got from the API in an excel file, sorted by the column "gruppe"
    data = pandas.DataFrame(filtered_response)
    data.sort_values(by="gruppe", inplace=True)

    data_dict_api = data.to_dict(orient="list")

    #----------------Updated code--------------------
    #getting the color code for each json item
    color_codes = []
    for item in data_dict_api["labelIds"]:
        if item is None:
            color_codes.append("")
        else:
            try:
                color_response = requests.get(f"{COLOR_API}{str(item)}").json()
                color_codes.append(color_response[0]['colorCode'].replace("#", ""))
            except IndexError:
                color_codes.append("")


    #---------------------Formating the 'hu' date, because we will need it later to determine the color of each row-------------------------

    #storing the date values into an array
    date_values = []
    for (columnName, columnData) in data.iteritems():
        if columnName == "hu":
            date_values.append(columnData.values)


    #converting the array into a string so we can replace the unneeded values
    dates_str = ' '.join([str(elem) for elem in date_values])


    #converting the string back to a list, whilst repalcing the unneeded values
    new_date_values = list(dates_str.replace("[", "").replace("]", "").replace("'", "").replace("\n", "").split(" "))
    today_date = current_date_iso_formatted.replace("-", "")

    final_date_values = []
    for date in new_date_values:
        new_date_value = datetime.strptime(date, '%Y-%m-%d').isoformat().split("T").pop(0).replace("-", "")
        final_date_values.append(int(today_date) - int(new_date_value))


    #function that is used to determine the color of each row in the table for the API data
    def apply_color(value) -> str:
        color_fill = str()
        #we converted the dates to integers and subtracted the current from the "hu" date
        #now we assign a color value base on the outcome of the subtraction

        if value < 300: #less than 3 months
            color_fill = "007500"
        elif value < 10000: ## between 3 months and a year
            color_fill = "FFA500"
        else: #more than a year
            color_fill = "b30000"

        return color_fill


    # --------------------------------Reading from the csv file--------------------------
    data_vehicles = pandas.read_csv('vehicles.csv', delimiter=";")
    data_vehicles.sort_values(by="gruppe", inplace=True)
    data_dict = data_vehicles.to_dict(orient="list")

    # getting the color codes of each labelId
    vehicle_color_codes = []
    for item in data_dict["labelIds"]:
        if math.isnan(item):
            vehicle_color_codes.append("")
        else:
            color_response = requests.get(f"{COLOR_API}{str(item)}").json()
            vehicle_color_codes.append(color_response[0]['colorCode'].replace("#", ""))



    #---------------Opening an Excel Writer to place the dataframes into an excel file------------------------------
    with pandas.ExcelWriter(f"vehicles_{current_date_iso_formatted}.xlsx", engine="openpyxl") as writer:
        #dropping the data that the user didn't request
        for key in filtered_response[0]:
            if key not in columns_to_print:
                data.drop(key, axis=1, inplace=True)

        #actually saving the data to an excel file
        sheet_name_api = "API Data Sheet"
        data.to_excel(writer, index=0, sheet_name=sheet_name_api)


        #saving the data from the csv in a separate sheet
        sheet_name_vehicles = "CSV Data Sheet"
        data_vehicles.to_excel(writer, index=0, sheet_name=sheet_name_vehicles)

    #-------------------Reading from the file and coloring the rows, if the user chose that option--------------------
    if c:
        #laoding a workbook and getting the needed worksheet
        wb = openpyxl.load_workbook(f"vehicles_{current_date_iso_formatted}.xlsx")
        ws = wb[sheet_name_api]

        #coloring each row of based on the "hu" value in it
        for (row, i) in zip(ws.iter_rows(min_row=2, max_col=len(data.columns), max_row=len(data) + 1), range(len(final_date_values))):
            for cell in row:
                cell.fill = PatternFill("solid", start_color=apply_color(final_date_values[i]))

        #-----------------Old way of coloring text (it colored all the cells in a row)
                # if "labelIds" in columns_to_print:
                #     if color_codes[i] == "":
                #         pass
                #     else:
                #         cell.font = Font(color=f"{color_codes[j]}")


    #-----------------New way of coloring the text in the column labelIds------------------
        #locating where the the labelIds column is
        if "labelIds" in columns_to_print:
            cell_needed = str()
            for col in ws.iter_cols(min_row=1, max_row=2, min_col=3):
                for cell in col:
                    if cell.value == "labelIds":
                        cell_needed = str(cell)

            #converting the cell object into a manegable string object
            cell_needed = cell_needed.replace("<Cell 'API Data Sheet'.", "").replace(">", "")
            split_cell = list(cell_needed).pop(0).lower()

            #converting the cell(which is a letter) to an integer
            #will probably work with only 26 columns (which is enough for the task) but it will give errors if the columns are more than 26
            column_number = ord(split_cell) - 96

            #coloring each cell if the labelIds column where a color code is resolved
            for col in ws.iter_cols(min_row=2, max_row=len(data) + 1, min_col=column_number, max_col=column_number):
                for (cell, i) in zip(col, range(len(color_codes))):
                    if color_codes[i] == "":
                        pass
                    else:
                        cell.font = Font(color=f"{color_codes[i]}")

        # getting the needed worksheet for the data from the csv file
        ws2 = wb[sheet_name_vehicles]


        #-----------------Old code. Colors the whole row---------------------
        #coloring rows where a color code for labelIds is provided and resolved
        # for (row, i) in zip(ws2.iter_rows(min_row=2, max_col=len(data_vehicles.columns), max_row=len(data_vehicles) + 1), range(len(vehicle_color_codes))):
        #     for cell in row:
        #         print(cell.value)
        #         if vehicle_color_codes[i] == "":
        #             pass
        #         else:
        #             cell.font = Font(color=f"{vehicle_color_codes[i]}")
        #---------------------------------------------------------------


        #New code
        #coloring the cell of labelIds if a color code is provided and resolved
        for col in ws2.iter_cols(min_row=2, max_row=len(data_vehicles) + 1, min_col=6, max_col=6):
            for (cell, i) in zip(col, range(len(vehicle_color_codes))):
                if vehicle_color_codes[i] == "":
                    pass
                else:
                    cell.font = Font(color=f"{vehicle_color_codes[i]}")



        #finally saving the excel file
        wb.save(f"vehicles_{current_date_iso_formatted}.xlsx")
        wb.close()

    print("[Finished in {:.2f}s]".format(time.time() - start_time))


if __name__ == "__main__":
    main()
